"""
Document Generation Tool
Generates PDF and Excel documents from structured data.
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List
import json

# PDF Generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("reportlab not installed. PDF generation will be limited.")

# Excel Generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel generation will be limited.")

logger = logging.getLogger(__name__)


class DocumentGenerator:
    """Generates professional PDF and Excel documents."""
    
    def __init__(self, output_dir: str = "./data/generated"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_pdf(
        self,
        title: str,
        content: str,
        filename: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Generate a professional PDF document with optimized formatting.
        
        Args:
            title: Document title
            content: Document content (supports markdown-like formatting)
            filename: Custom filename (auto-generated if not provided)
            metadata: Additional metadata (author, subject, etc.)
        
        Returns:
            Dict with file_path and metadata
        """
        if not REPORTLAB_AVAILABLE:
            return self._generate_text_pdf_fallback(title, content, filename)
        
        try:
            # Generate filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{title.replace(' ', '_').replace(':', '')}_{timestamp}.pdf"
            
            file_path = self.output_dir / filename
            
            # Create PDF with better margins
            doc = SimpleDocTemplate(
                str(file_path),
                pagesize=letter,
                rightMargin=60,
                leftMargin=60,
                topMargin=60,
                bottomMargin=50
            )
            
            # Enhanced Styles
            styles = getSampleStyleSheet()
            
            # Cover title style
            cover_title_style = ParagraphStyle(
                'CoverTitle',
                parent=styles['Heading1'],
                fontSize=32,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=20,
                spaceBefore=100,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            # Subtitle style
            subtitle_style = ParagraphStyle(
                'Subtitle',
                parent=styles['Normal'],
                fontSize=14,
                textColor=colors.HexColor('#6B7280'),
                spaceAfter=30,
                alignment=TA_CENTER,
                fontName='Helvetica'
            )
            
            # Section heading style
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=14,
                spaceBefore=20,
                fontName='Helvetica-Bold',
                borderPadding=8,
                leftIndent=0
            )
            
            # Subheading style
            subheading_style = ParagraphStyle(
                'Subheading',
                parent=styles['Heading3'],
                fontSize=13,
                textColor=colors.HexColor('#374151'),
                spaceAfter=10,
                spaceBefore=14,
                fontName='Helvetica-Bold'
            )
            
            # Body text style with better spacing and line height
            body_style = ParagraphStyle(
                'CustomBody',
                parent=styles['BodyText'],
                fontSize=11,
                leading=18,  # Line height
                spaceAfter=10,
                alignment=TA_JUSTIFY,
                fontName='Helvetica',
                textColor=colors.HexColor('#1F2937')
            )
            
            # List item style
            list_style = ParagraphStyle(
                'ListItem',
                parent=body_style,
                leftIndent=20,
                bulletIndent=5,
                spaceAfter=6,
                spaceBefore=2
            )
            
            # Build document
            story = []
            
            # === COVER PAGE ===
            story.append(Spacer(1, 1.5 * inch))
            story.append(Paragraph(title, cover_title_style))
            
            # Subtitle with metadata
            if metadata and metadata.get('subject'):
                story.append(Paragraph(f"Informe sobre: {metadata['subject']}", subtitle_style))
            
            story.append(Spacer(1, 0.5 * inch))
            
            # Metadata box
            meta_lines = [
                f"<b>Fecha:</b> {datetime.now().strftime('%d de %B de %Y')}",
                f"<b>Hora:</b> {datetime.now().strftime('%H:%M')}"
            ]
            if metadata:
                if metadata.get('author'):
                    meta_lines.append(f"<b>Generado por:</b> {metadata['author']}")
                if metadata.get('sources'):
                    meta_lines.append(f"<b>Fuentes:</b> {metadata['sources']}")
            
            meta_box_style = ParagraphStyle(
                'MetaBox',
                parent=styles['Normal'],
                fontSize=10,
                leading=16,
                alignment=TA_CENTER,
                textColor=colors.HexColor('#6B7280')
            )
            
            for line in meta_lines:
                story.append(Paragraph(line, meta_box_style))
                story.append(Spacer(1, 0.05 * inch))
            
            # Page break after cover
            story.append(PageBreak())
            
            # === CONTENT PROCESSING ===
            # Clean and normalize content
            content = content.strip()
            
            # DEBUG: Log original content
            logger.info(f"🔍 Original content preview (first 300 chars): {content[:300]}")
            logger.info(f"🔍 Newline count in original: single=\\n:{content.count(chr(10))}, words={len(content.split())}")
            
            # First, intelligently join lines that belong together
            # This fixes PDFs where each word is on a separate line
            content = self._normalize_text_flow(content)
            
            # DEBUG: Log normalized content
            logger.info(f"🔍 Normalized content preview (first 300 chars): {content[:300]}")
            
            # Split by double newlines for paragraphs
            sections = content.split('\n\n')
            
            for section in sections:
                section = section.strip()
                if not section:
                    continue
                
                # Check for table markers
                if section.startswith('TABLE_START'):
                    table_lines = section.split('\n')[1:]
                    table_data = [['<b>Campo</b>', '<b>Valor</b>']]
                    
                    for line in table_lines:
                        if line.strip() == 'TABLE_END':
                            break
                        if '|' in line:
                            parts = line.split('|', 1)
                            if len(parts) == 2:
                                table_data.append([parts[0].strip(), parts[1].strip()])
                    
                    if len(table_data) > 1:
                        t = Table(table_data, colWidths=[2.2*inch, 4.3*inch])
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 11),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                            ('TOPPADDING', (0, 0), (-1, 0), 10),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                            ('FONTSIZE', (0, 1), (-1, -1), 10),
                            ('TOPPADDING', (0, 1), (-1, -1), 8),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                            ('LEFTPADDING', (0, 0), (-1, -1), 10),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white])
                        ]))
                        story.append(Spacer(1, 0.15 * inch))
                        story.append(t)
                        story.append(Spacer(1, 0.25 * inch))
                    continue
                
                # Section headings (##)
                if section.startswith('## '):
                    heading_text = section[3:].strip()
                    story.append(Paragraph(heading_text, heading_style))
                    continue
                
                # Main headings (#)
                if section.startswith('# '):
                    main_heading_text = section[2:].strip()
                    story.append(Paragraph(main_heading_text, cover_title_style))
                    story.append(Spacer(1, 0.2 * inch))
                    continue
                
                # Subheadings (###)
                if section.startswith('### '):
                    subhead_text = section[4:].strip()
                    story.append(Paragraph(subhead_text, subheading_style))
                    continue
                
                # Handle lists (detect multiple lines starting with - or *)
                lines = section.split('\n')
                if len(lines) > 1 and all(line.strip().startswith(('- ', '* ', '• ')) for line in lines if line.strip()):
                    # It's a list
                    for line in lines:
                        line = line.strip()
                        if not line:
                            continue
                        # Remove bullet character
                        if line.startswith(('- ', '* ')):
                            line_text = line[2:].strip()
                        elif line.startswith('• '):
                            line_text = line[2:].strip()
                        else:
                            line_text = line
                        
                        # Apply formatting
                        line_text = self._apply_inline_formatting(line_text)
                        story.append(Paragraph(f"• {line_text}", list_style))
                    story.append(Spacer(1, 0.15 * inch))
                    continue
                
                # Normal paragraph - process text
                processed_text = self._apply_inline_formatting(section)
                
                if processed_text:
                    story.append(Paragraph(processed_text, body_style))
                    story.append(Spacer(1, 0.1 * inch))
            
            # === PAGE FOOTER ===
            def add_page_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.HexColor('#9CA3AF'))
                
                # Left footer
                canvas.drawString(60, 30, f"ServiBot Report | {datetime.now().strftime('%d/%m/%Y')}")
                
                # Right footer (page number)
                if doc.page > 1:  # Don't show page number on cover
                    canvas.drawRightString(letter[0] - 60, 30, f"Página {doc.page - 1}")
                
                # Top border line
                canvas.setStrokeColor(colors.HexColor('#E5E7EB'))
                canvas.setLineWidth(0.5)
                canvas.line(60, letter[1] - 40, letter[0] - 60, letter[1] - 40)
                
                canvas.restoreState()
            
            doc.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
            
            logger.info(f"✅ Generated PDF: {file_path}")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "size_bytes": file_path.stat().st_size,
                "format": "pdf"
            }
            
        except Exception as e:
            logger.error(f"❌ PDF generation failed: {e}")
            logger.exception(e)
            return {"success": False, "error": str(e)}
    
    def _normalize_text_flow(self, text: str) -> str:
        """
        Normalize text flow by intelligently joining lines.
        Fixes PDFs where each word appears on a separate line.
        """
        lines = text.split('\n')
        result = []
        buffer = []
        empty_line_count = 0
        
        logger.info(f"🔍 _normalize_text_flow input: {len(lines)} lines")
        logger.info(f"🔍 First 10 lines: {lines[:10]}")
        
        for i, line in enumerate(lines):
            stripped = line.strip()
            
            # Count consecutive empty lines
            if not stripped:
                empty_line_count += 1
                # Only treat as paragraph break if we have 2+ empty lines OR significant buffer
                if empty_line_count >= 2 and buffer:
                    result.append(' '.join(buffer))
                    buffer = []
                    if i < len(lines) - 1:  # Don't add trailing empty
                        result.append('')
                    empty_line_count = 0
                continue
            
            # Reset empty line counter
            empty_line_count = 0
            
            # Headings, tables, lists - flush buffer and keep separate
            if stripped.startswith(('#', 'TABLE_START', 'TABLE_END')):
                if buffer:
                    result.append(' '.join(buffer))
                    buffer = []
                result.append(stripped)
                continue
            
            # List items - flush buffer and keep separate
            if stripped.startswith(('-', '*', '•')) and len(stripped) > 1:
                if buffer:
                    result.append(' '.join(buffer))
                    buffer = []
                result.append(stripped)
                continue
            
            # Accumulate words/phrases
            buffer.append(stripped)
            
            # Close buffer on sentence endings
            if stripped.endswith(('.', '!', '?', ':', ';')):
                result.append(' '.join(buffer))
                buffer = []
        
        # Flush remaining
        if buffer:
            result.append(' '.join(buffer))
        
        output = '\n\n'.join(result)
        logger.info(f"🔍 _normalize_text_flow output length: {len(output)} chars")
        logger.info(f"🔍 Output first 500: {output[:500]}")
        
        return output
    
    def _process_paragraph_text(self, text: str) -> str:
        """
        Process paragraph text to join lines intelligently.
        Fixes the issue where each word is on a separate line.
        """
        lines = text.split('\n')
        processed_lines = []
        current_para = []
        
        for line in lines:
            line = line.strip()
            if not line:
                if current_para:
                    processed_lines.append(' '.join(current_para))
                    current_para = []
                continue
            
            # If line ends with punctuation or is a heading, it's a complete thought
            if line.endswith(('.', '!', '?', ':', ';')) or line.startswith(('#', '-', '*', '•')):
                current_para.append(line)
                processed_lines.append(' '.join(current_para))
                current_para = []
            else:
                # Continue building current paragraph
                current_para.append(line)
        
        # Add remaining
        if current_para:
            processed_lines.append(' '.join(current_para))
        
        return ' '.join(processed_lines)
    
    def _apply_inline_formatting(self, text: str) -> str:
        """Apply inline formatting like bold, italic, etc."""
        import re
        # Replace **text** with <b>text</b>
        text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
        # Replace *text* with <i>text</i> (but not if already in bold)
        text = re.sub(r'(?<!</b>)\*(?!<b>)(.+?)\*', r'<i>\1</i>', text)
        return text
    
    def generate_excel(
        self,
        title: str,
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = "Sheet1"
    ) -> Dict[str, Any]:
        """
        Generate an Excel spreadsheet.
        
        Args:
            title: Spreadsheet title
            data: List of dictionaries (each dict is a row)
            filename: Custom filename
            sheet_name: Name of the worksheet
        
        Returns:
            Dict with file_path and metadata
        """
        if not OPENPYXL_AVAILABLE:
            return self._generate_csv_fallback(title, data, filename)
        
        try:
            # Generate filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{title.replace(' ', '_')}_{timestamp}.xlsx"
            
            file_path = self.output_dir / filename
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Styles
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if not data:
                return {"success": False, "error": "No data provided"}
            
            # Headers
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header.replace('_', ' ').title()
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            for row_num, row_data in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = row_data.get(header, '')
                    cell.border = border
                    
                    # Auto-format numbers and dates
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(header))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save
            wb.save(file_path)
            
            logger.info(f"✅ Generated Excel: {file_path}")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "size_bytes": file_path.stat().st_size,
                "format": "xlsx",
                "rows": len(data)
            }
            
        except Exception as e:
            logger.error(f"❌ Excel generation failed: {e}")
            return {"success": False, "error": str(e)}
    
    def _generate_text_pdf_fallback(self, title: str, content: str, filename: Optional[str]) -> Dict[str, Any]:
        """Fallback: generate text file when reportlab not available."""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{title.replace(' ', '_')}_{timestamp}.txt"
            
            file_path = self.output_dir / filename
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"{title}\n")
                f.write("=" * len(title) + "\n\n")
                f.write(content)
            
            logger.warning(f"⚠️ Generated text fallback (reportlab not available): {file_path}")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "size_bytes": file_path.stat().st_size,
                "format": "txt",
                "fallback": True
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _generate_csv_fallback(self, title: str, data: List[Dict], filename: Optional[str]) -> Dict[str, Any]:
        """Fallback: generate CSV when openpyxl not available."""
        try:
            import csv
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{title.replace(' ', '_')}_{timestamp}.csv"
            
            file_path = self.output_dir / filename
            
            if not data:
                return {"success": False, "error": "No data provided"}
            
            headers = list(data[0].keys())
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)
            
            logger.warning(f"⚠️ Generated CSV fallback (openpyxl not available): {file_path}")
            
            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "size_bytes": file_path.stat().st_size,
                "format": "csv",
                "fallback": True,
                "rows": len(data)
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


# Singleton instance
_generator: Optional[DocumentGenerator] = None


def get_document_generator() -> DocumentGenerator:
    """Get or create DocumentGenerator singleton."""
    global _generator
    if _generator is None:
        _generator = DocumentGenerator()
    return _generator
