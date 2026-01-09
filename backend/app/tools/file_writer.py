"""
File Writer Tool - Generate PDF and Excel files
Creates formatted documents from structured data
"""
from typing import Dict, Any, List, Optional
import logging
from pathlib import Path
from datetime import datetime
import os

logger = logging.getLogger(__name__)


class FileWriterTool:
    """Tool for generating PDF and Excel files from structured data."""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize the file writer tool.
        
        Args:
            output_dir: Directory to save generated files (defaults to ./data/generated)
        """
        self.output_dir = output_dir or os.path.join(os.getcwd(), "data", "generated")
        os.makedirs(self.output_dir, exist_ok=True)
        logger.info(f"FileWriterTool initialized with output dir: {self.output_dir}")
    
    def generate_pdf(
        self,
        title: str,
        content: str,
        filename: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Generate a PDF document.
        
        Args:
            title: Document title
            content: Main content (supports paragraphs separated by \\n\\n)
            filename: Optional custom filename (auto-generated if not provided)
            metadata: Optional metadata dict with author, subject, etc.
            
        Returns:
            Dict with status, file_path, and filename
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            return {
                "status": "error",
                "message": "reportlab library not installed"
            }
        
        try:
            # Generate filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{title[:30].replace(' ', '_')}.pdf"
            
            if not filename.endswith('.pdf'):
                filename += '.pdf'
            
            file_path = os.path.join(self.output_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Container for flowables
            story = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor='#2C3E50',
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            body_style = ParagraphStyle(
                'CustomBody',
                parent=styles['BodyText'],
                fontSize=12,
                leading=16,
                alignment=TA_JUSTIFY,
                spaceAfter=12
            )
            
            # Add title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.2 * inch))
            
            # Add metadata if provided
            if metadata:
                meta_style = ParagraphStyle(
                    'Metadata',
                    parent=styles['Normal'],
                    fontSize=10,
                    textColor='#7F8C8D',
                    spaceAfter=20
                )
                meta_text = " | ".join([f"<b>{k}:</b> {v}" for k, v in metadata.items()])
                story.append(Paragraph(meta_text, meta_style))
                story.append(Spacer(1, 0.3 * inch))
            
            # Add content (split by paragraphs)
            paragraphs = content.split('\n\n')
            for para in paragraphs:
                if para.strip():
                    story.append(Paragraph(para.strip(), body_style))
                    story.append(Spacer(1, 0.1 * inch))
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF generated successfully: {file_path}")
            return {
                "status": "success",
                "file_path": file_path,
                "filename": filename,
                "message": f"PDF created: {filename}"
            }
            
        except Exception as e:
            logger.exception(f"Error generating PDF: {e}")
            return {
                "status": "error",
                "message": f"Failed to generate PDF: {str(e)}"
            }
    
    def generate_excel(
        self,
        filename: str,
        sheets: Dict[str, List[List[Any]]],
        headers: Optional[Dict[str, List[str]]] = None
    ) -> Dict[str, Any]:
        """
        Generate an Excel file with multiple sheets.
        
        Args:
            filename: Output filename
            sheets: Dict mapping sheet names to 2D lists of data
            headers: Optional dict mapping sheet names to header row lists
            
        Returns:
            Dict with status, file_path, and filename
            
        Example:
            sheets = {
                "Sales": [
                    ["Product", "Q1", "Q2", "Q3", "Q4"],
                    ["Widget A", 100, 150, 200, 180],
                    ["Widget B", 80, 90, 110, 120]
                ]
            }
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return {
                "status": "error",
                "message": "openpyxl library not installed"
            }
        
        try:
            # Generate filename
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            file_path = os.path.join(self.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name, data in sheets.items():
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers if provided
                if headers and sheet_name in headers:
                    header_row = headers[sheet_name]
                    for col_idx, header in enumerate(header_row, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    start_row = 2
                else:
                    start_row = 1
                
                # Add data
                for row_idx, row_data in enumerate(data, start_row):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Auto-adjust column width
                        column_letter = cell.column_letter
                        if ws.column_dimensions[column_letter].width < len(str(value)) + 2:
                            ws.column_dimensions[column_letter].width = len(str(value)) + 2
            
            # Save workbook
            wb.save(file_path)
            
            logger.info(f"Excel file generated successfully: {file_path}")
            return {
                "status": "success",
                "file_path": file_path,
                "filename": filename,
                "sheets": list(sheets.keys()),
                "message": f"Excel created: {filename} with {len(sheets)} sheet(s)"
            }
            
        except Exception as e:
            logger.exception(f"Error generating Excel: {e}")
            return {
                "status": "error",
                "message": f"Failed to generate Excel: {str(e)}"
            }
    
    def generate_report(
        self,
        report_type: str,
        data: Dict[str, Any],
        format: str = "pdf"
    ) -> Dict[str, Any]:
        """
        Generate a formatted report (PDF or Excel).
        
        Args:
            report_type: Type of report (summary, detailed, custom)
            data: Data to include in the report
            format: Output format ("pdf" or "excel")
            
        Returns:
            Dict with status and file information
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format.lower() == "pdf":
            title = data.get("title", f"{report_type.title()} Report")
            content = data.get("content", "")
            
            if not content and "items" in data:
                # Auto-format items as content
                items = data["items"]
                content = "\n\n".join([
                    f"• {item}" if isinstance(item, str) else f"• {item.get('name', 'Item')}: {item.get('value', '')}"
                    for item in items
                ])
            
            metadata = {
                "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Type": report_type.title(),
                "Author": "ServiBot"
            }
            
            filename = f"{timestamp}_{report_type}_report.pdf"
            return self.generate_pdf(title, content, filename, metadata)
        
        elif format.lower() == "excel":
            sheets = {}
            
            if "sheets" in data:
                sheets = data["sheets"]
            elif "items" in data:
                # Auto-convert items to sheet
                sheets["Report Data"] = [[k, v] for k, v in data["items"].items()] if isinstance(data["items"], dict) else [data["items"]]
            
            headers = data.get("headers", {})
            filename = f"{timestamp}_{report_type}_report.xlsx"
            return self.generate_excel(filename, sheets, headers)
        
        else:
            return {
                "status": "error",
                "message": f"Unsupported format: {format}. Use 'pdf' or 'excel'"
            }


# Singleton instance
_file_writer_instance = None

def get_file_writer() -> FileWriterTool:
    """Get or create the file writer tool singleton."""
    global _file_writer_instance
    if _file_writer_instance is None:
        _file_writer_instance = FileWriterTool()
    return _file_writer_instance
